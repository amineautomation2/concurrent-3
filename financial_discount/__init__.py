from utils import delay, save_xlsx
from worker import get_data_by_worker_id, get_xlsx_data, write_csv_by_id
from math import ceil
import curl_cffi
import json
import io
import openpyxl
from time import sleep
from random import uniform
from pypdf import PdfReader
from re import findall
from utils import get_xlsx_filepath, get_random_user_agent, fetch_with_backoff


def isin_from_text(text: str) -> str:
    isin_pattern = r"[A-Z]{2}[A-Z0-9]{9}[0-9]"
    isin = findall(isin_pattern, text)
    if len(isin) > 0:
        return isin[0]
    return ""


def isin_from_pdf(url: str) -> str:
    cookies = {
        'safariCookie': '1',
        'ASP.NET_SessionId': 'ffhz34mnbxiumr3rlfodsjaw',
    }

    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-US,en;q=0.9',
        'cache-control': 'max-age=0',
        'priority': 'u=0, i',
        'sec-ch-ua': '"Not(A:Brand";v="8", "Chromium";v="144"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
    }

    headers.update(get_random_user_agent())
    isin = ""
    if len(url) == 0:
        return isin

    response = fetch_with_backoff(url, headers=headers, cookies=cookies)
    if response:
        if response.content:
            try:
                pdf_bytes = io.BytesIO(response.content)
                reader = PdfReader(pdf_bytes)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() or ""
            except Exception as e:
                print(f"[{url}]isin_from_pdf: ", e)
                return isin
            return isin_from_text(text)
    return isin


def get_total_funds() -> int:
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'en-US,en;q=0.9',
        # 'Accept-Encoding': 'gzip, deflate, br, zstd',
        'Origin': 'https://investorhub.financialexpress.net',
        'Sec-GPC': '1',
        'Connection': 'keep-alive',
        'Referer': 'https://investorhub.financialexpress.net/',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'cross-site',
        # Requests doesn't support trailers
        # 'TE': 'trailers',
    }

    ua = get_random_user_agent()
    headers.update(ua)
    params = {
        'jsonString': '{"FilteringOptions":{"undefined":0,"RangeId":null,"RangeName":"","CategoryId":null,"Category2Id":null,"PriipProductCode":null,"DefaultCategoryId":null,"DefaultCategory2Id":null,"ForSaleIn":null,"ShowMainUnits":false,"MPCategoryCode":null},"ProjectName":"fdd","LanguageCode":"en-gb","UserType":"","Region":"","LanguageId":"1","LocaleId":"1","Theme":"fdd","SortingStyle":"1","PageNo":1,"PageSize":25,"OrderBy":"UnitName:init","IsAscOrder":true,"OverrideDocumentCountryCode":null,"ToolId":"1","PrefetchPages":80,"PrefetchPageStart":1,"OverridenThemeName":"fdd","ForSaleIn":"","ValidateFeResearchAccess":false,"HasFeResearchFullAccess":false,"EnableSedolSearch":"false","GrsProjectId":"17200043","ShowMainUnitExpansion":false,"UseCombinedOngoingChargeTER":false}',
    }

    response = curl_cffi.get(
        'https://digitalfundservice.feprecisionplus.com/FundDataService.svc/GetRowIdList',
        params=params,
        headers=headers,
        impersonate="chrome"
    )
    if response.status_code != 200:
        return 0

    total = response.json()["TotalRows"]
    return int(total)


def get_rows_id(begin: int, end: int) -> str:
    rows = []
    for i in range(begin, end + 1):
        rows.append(f"{i}")
    return ",".join(rows)


def get_page_data(nb_page: int, page_size: int, rows_id: str) -> list[dict]:
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'en-US,en;q=0.9',
        # 'Accept-Encoding': 'gzip, deflate, br, zstd',
        'Origin': 'https://investorhub.financialexpress.net',
        'Sec-GPC': '1',
        'Connection': 'keep-alive',
        'Referer': 'https://investorhub.financialexpress.net/',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'cross-site',
        'Priority': 'u=0',
    }

    ua = get_random_user_agent()
    headers.update(ua)

    params = '{"FilteringOptions":{"undefined":0,"RangeId":null,"RangeName":"","CategoryId":null,"Category2Id":null,"PriipProductCode":null,"DefaultCategoryId":null,"DefaultCategory2Id":null,"ForSaleIn":null,"ShowMainUnits":false,"MPCategoryCode":null},"ProjectName":"fdd","LanguageCode":"en-gb","UserType":"","Region":"","LanguageId":"1","LocaleId":"1","Theme":"fdd","SortingStyle":"1","PageNo":2,"PageSize":25,"OrderBy":"UnitName:init","IsAscOrder":true,"OverrideDocumentCountryCode":null,"ToolId":"1","PrefetchPages":80,"PrefetchPageStart":1,"OverridenThemeName":"fdd","ForSaleIn":"","ValidateFeResearchAccess":false,"HasFeResearchFullAccess":false,"EnableSedolSearch":"false","RowCount":3600,"RowIDs":"26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50"}'

    payload = json.loads(params)
    pageInfo = {
        "PageNo": nb_page,
        "PageSize": page_size,
        "RowIDs": f"{rows_id}",
    }
    payload.update(pageInfo)
    params = {
        'jsonString': json.dumps(payload),
    }
    response = curl_cffi.get(
        'https://digitalfundservice.feprecisionplus.com/FundDataService.svc/GetUnitList',
        params=params,
        headers=headers,
        impersonate='chrome'
    )
    if response.status_code != 200:
        print(response)
        return []

    data = response.json()
    decode = json.loads(data)
    return decode["DataList"]
    # with open(f"financial_{nb_page}.json", "w+") as f:
    #    json.dump(decode, f, indent=4)


def get_funds_url() -> list[dict]:
    base_url = "https://investorhub.financialexpress.net"
    start = 1
    page_size = 25
    total_funds = get_total_funds()

    print(f"[financial discount] total funds = {total_funds}")
    if total_funds == 0:
        raise Exception(f"financial discount: total_funds = {total_funds}")

    total_pages = ceil(total_funds / page_size)
    end = page_size * start
    funds_url = []
    for current_page in range(start, total_pages+1):
        print(f"page [{current_page}/{total_pages}]")
        rows = get_rows_id(start, end)
        page_funds_json = get_page_data(current_page, page_size, rows)

        for fund in page_funds_json:
            name = fund["FundInfo"]["Name"]
            url = f'{base_url}{fund["FundInfo"]["FactsheetPdfLink"]}'
            # print(fund_data.get("name"))
            funds_url.append(dict(name=name, url=url))
        start += page_size
        end += page_size
        sleep(uniform(1, 2))
    return funds_url


def extract_isin(fund: dict) -> str | None:
    url = fund.get("url")
    if url:
        return isin_from_pdf(url)


def financial_discount_runner(id_worker: int, max_worker: int) -> None:
    out_xlsx = get_xlsx_filepath("financial_discount.xlsx")
    data = get_xlsx_data(out_xlsx, "Funds")
    funds_per_worker = get_data_by_worker_id(id_worker, max_worker, data[:5])
    for fund in funds_per_worker:
        url = fund.get("url")
        if url:
            isin = isin_from_pdf(url)
            fund.update(dict(isin=isin))
        delay(2, 4)

    csv = f"financial_discount_{id_worker}.csv"
    fields = ["index", "name", "isin", "url"]
    write_csv_by_id(csv, funds_per_worker, fields)


def get_financial_url(xlsx: str) -> None:
    urls = get_funds_url()
    save_xlsx(
        xlsx_out=xlsx,
        funds=urls,
        cols=["name", "isin", "url"],
        sheet="Funds",
    )


def financial_write_xlsx(file_xlsx: str, data: list[dict]):
    wb = openpyxl.load_workbook(file_xlsx)
    ws = wb["Funds"]
    final = []

    for fund in data:
        if not fund.get("isin"):
            url = fund.get("url")
            if url:
                print(f"extracting from {url}")
                isin = isin_from_pdf(url)
                fund.update(dict(isin=isin))
        final.append(fund)
        sleep(uniform(3, 10))

    i = 2
    for row in final:
        ws.cell(i, 1, row["name"])
        ws.cell(i, 2, row["isin"])
        c = ws.cell(i, 3, row["url"])
        c.hyperlink = row["url"]
        c.style = "Hyperlink"
        i += 1
    wb.save(file_xlsx)

    for sheet in wb:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # type: ignore

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_xlsx)
    wb.close()
