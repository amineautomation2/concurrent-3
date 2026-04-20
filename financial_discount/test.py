from openpyxl import load_workbook
from pypdf.errors import EmptyFileError
from pypdf import PdfReader
from re import findall
import utils
from random import uniform
from os.path import join, dirname, basename
from pathlib import Path
import openpyxl
from utils import fetch_with_backoff
from selenium import webdriver
import base64
from time import sleep
from utils import setup_driver
from pprint import pprint
import json
from utils import get_xlsx_filepath
from os.path import join
from os import chdir
from os.path import dirname
import requests
from utils import get_random_user_agent
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def create_mutual_fund_spreadsheet(data, filepath):
    """
    Create Excel spreadsheet with Mutual Fund data.

    Args:
        data: List of dictionaries containing Name, Open, and ISIN
        filepath: Output file path for the Excel file
    """
    # Create new workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Mutual Fund"

    # Style the header row
    header_font = Font(bold=True, italic=True, size=13, color='000000')
    header_fill = PatternFill(start_color='4472C4',
                              end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Write headers
    headers = ['Name', 'Open', 'ISIN']
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    # for row in data:
    #    sheet.append([
    #        row.get('Name'),
    #        row.get('Open'),
    #        row.get('ISIN')
    #    ])

    for column in sheet.columns:  # type: ignore
        max_length = 0
        column_letter = column[0].column_letter  # type: ignore

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max_length + 2
        # type: ignore
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    workbook.save(filepath)
    workbook.close()


# mutual_funds = [
#    {'Name': 'Vanguard 500 Index', 'Open': 425.32, 'ISIN': 'US9229083632'},
#    {'Name': 'Fidelity Total Market', 'Open': 115.67, 'ISIN': 'US3160928030'},
#    {'Name': 'PIMCO Income Fund', 'Open': 11.85, 'ISIN': 'US72201R8824'}
# ]

# create_mutual_fund_spreadsheet(mutual_funds, 'mutual_funds.xlsx')
# print('Spreadsheet created successfully!')


# mylogger = logging.getLogger(__name__)
# logging_config = {
#    "version": 1,
#    "disable_existing_loggers": False,
#    "formatters": {
#        "simple": {
#            "format": "%(levelname)s: %(message)s",
#        }
#    },
#    "handlers": {
#        "stderr": {
#            "class": "logging.StreamHandler",
#            "level": "WARNING",
#            "formatter": "simple",
#            "stream": "ext://sys.stderr",
#        },
#        "file": {
#            "class": "logging.handlers.RotatingFileHandler",
#            "level": "DEBUG",
#            "formatter": "simple",
#            "filename": "logs/app.log",
#            "maxBytes": 10000,
#            "backupCount": 3,
#        }
#    },
#    "loggers": {
#        "root": {
#            "level": "DEBUG",
#            "handlers": ["stdout"]
#        }
#    },
# }
#
# logging.config.dictConfig(config=logging_config)
# mylogger.addHandler(logging.StreamHandler())
# mylogger.debug("debug message")
# mylogger.info("info message")
# mylogger.warning("warning message")
# mylogger.error("error message")
# mylogger.critical("critical message")
#
#
# def calc():
#    try:
#        s = 1 / 0
#    except:
#        raise Exception(ZeroDivisionError)
#


# print(get_xlsx_filepath("halifax.xlsx"))


def search_by_name_instagram(search_query: str):
    cookies = {
        'csrftoken': 'T3iSCXhmgAH2cPhmYpGYjW',
        'datr': 'vNN_aehDrxaSFW1EvXy6ZgMt',
        'ig_did': 'AE3E14B2-3927-40C1-A940-5A9095A742A0',
        'mid': 'aX_TvAALAAEva4iT6DkyYfxmi9T3',
        'ig_nrcb': '1',
        'ds_user_id': '66263675889',
        'sessionid': '66263675889%3APU6zzOQ2JcoGLg%3A1%3AAYipeUiTgHBU0DJdVQJJQSNreqVm28RrVnkKDwVPvOA',
        'ps_l': '1',
        'ps_n': '1',
        'wd': '813x927',
        'rur': '"CLN\\05466263675889\\0541805580432:01feaa8b16bb5e889649ff215f8c3dcb5c02071aacae474974953ebabd314a12dae15f10"',
    }

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:148.0) Gecko/20100101 Firefox/148.0',
        'Accept': '*/*',
        'Accept-Language': 'en-US,en;q=0.9',
        # 'Accept-Encoding': 'gzip, deflate, br, zstd',
        'Content-Type': 'application/x-www-form-urlencoded',
        'X-FB-Friendly-Name': 'PolarisSearchBoxRefetchableQuery',
        'X-CSRFToken': 'T3iSCXhmgAH2cPhmYpGYjW',
        'X-IG-App-ID': '936619743392459',
        'X-BLOKS-VERSION-ID': '9c0aa96c08c5b24220ee33094940e011645902f00d10e21e03b027ede1dc2735',
        'X-Root-Field-Name': 'xdt_api__v1__fbsearch__topsearch_connection',
        'X-FB-LSD': 'pD0B-QELbLVH3Ixu_JX8Cs',
        'X-ASBD-ID': '359341',
        'Origin': 'https://www.instagram.com',
        'Sec-GPC': '1',
        'Connection': 'keep-alive',
        'Referer': 'https://www.instagram.com/boutique.casavelo/',
        # 'Cookie': 'csrftoken=T3iSCXhmgAH2cPhmYpGYjW; datr=vNN_aehDrxaSFW1EvXy6ZgMt; ig_did=AE3E14B2-3927-40C1-A940-5A9095A742A0; mid=aX_TvAALAAEva4iT6DkyYfxmi9T3; ig_nrcb=1; ds_user_id=66263675889; sessionid=66263675889%3APU6zzOQ2JcoGLg%3A1%3AAYipeUiTgHBU0DJdVQJJQSNreqVm28RrVnkKDwVPvOA; ps_l=1; ps_n=1; wd=813x927; rur="CLN\\05466263675889\\0541805580432:01feaa8b16bb5e889649ff215f8c3dcb5c02071aacae474974953ebabd314a12dae15f10"',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        # Requests doesn't support trailers
        # 'TE': 'trailers',
    }
    query = f'{{"data":{{"context":"blended","include_reel":"true","query":"{search_query}","rank_token":"","search_session_id":"97985cb9-a401-43ac-a65d-51d5341dcbd2","search_surface":"web_top_search"}},"hasQuery":true}}'
    data = {
        'av': '17841466356672959',
        '__d': 'www',
        '__user': '0',
        '__a': '1',
        '__req': '1g',
        '__hs': '20532.HYP:instagram_web_pkg.2.1...0',
        'dpr': '1',
        '__ccg': 'UNKNOWN',
        '__rev': '1035610417',
        '__s': '3i9ppm:yo9n82:hhx4g7',
        '__hsi': '7619460331507933713',
        '__dyn': '7xeUjG1mxu1syaxG4Vp41twpUnwgU7SbzEdF8aUco2qwJyEiw9-1DwUx609vCwjE1EEc87m0yE462mcw5Mx62G5UswoEcE7O2l0Fwqo31w9a9wlo8od8-U2exi4UaEW2G0AEco5G1Wxfxm16wUwxwl8vwww51wLyESE7i3vwDwHg2ZwrUK2K2WE5B08-269wr86C1mgcEed6hEhK2OubK5V89FbxG1oxe6U5q0EoKmUhw4rwxxCaCwHwi84q2i1cw',
        '__csr': 'h75MqhBP8Y64aNiN4lf5l_tSOlsLtjn98Sx3hSKC4GVW88FvH-ytv9FSivrRU_ExSuiK6pQmbFkQ-rjAl9eRAFamrKmrQiqHiCKHFmS8Gpaji-Z4t2iVEjh9RF3BK4kDHrAkFd1V2ELhF8yl7yGQpBxmaWHjLKQiijVQXCCyEXDJAHg-leYyul5BGJ4h4mUCmlt4xnDzaWzkqqeAghLyEHUCqiGJlguGcUlzoW3y1wGhJsEfU-9w0o8U08aE0wG00VIapGcYgt0BwLwbm5-14AAxq26u2i0tai0VaxS1sa1LAYF8kwTw_AwroK0Uiw5nweP4i0Cx9woEqqxh2oQnc1aw2ny04qwiE6-0gq3sy8po1681kQtxgy80d_wn9o8AV4mzz8o8q2J1u5kc4a0xo1x86G0Kswpwmk7Qm4swuxomahe48a45U27w5IwgEwg1pwDAzUadx7ixhGcaGwJG3Za8rwq6220y80k8Ax21awAwsy28vDw8Gaw4QwlU5-9zrx5o--q4Q02zq5bnGy3Ve08tw1dO0qe0SE0TK5BwSw5Kw_U0U2Pk089J4g6W06xU0zO0xE1S-0qS0XoWdw6Sxim45Lg',
        '__hsdp': 'gtA1WwhwBq8I8MZbbcjMRp9R7cx3j4PT3YqVdPj8fFMSwa5rGyOhr2Fn8yAi88Wkm2314Qxsudg-9c54ehUJ5wt8G98m58S3hJBHxgw-6Ud6Q1gxh8VUCm1pgcUy4EqwjE9obk6EigWi2W5Eng5G7EdQaCAG1ewqoy0p20mG0oS0CU3ow5yg0yO11w6dwxwc-6A2e0PE3jU1wUcE2nw3e81ak0yU6K1dc3u19waG0OEpw',
        '__hblp': '0uU5qUfEdQ5UpWyAGyk689EfEjwRU4C4VU8U4Km5FQVV9UycglynyE4ly45eciK12DykaUgV8MwSXAggBGeyyF3oGt388pEvyki8xquUgByFXgcRBy8x3UB1yEWuiEjVeWzVp8a4czoN28Ki5UR7xqcym6o-4o-9AAxem4o8GgGqiEeojg98dUNKE4G1hxKex63q2m1pw5GwcuE32w8m5o7S1qxy1dwjo5qUK14a5o563Gqq222e7ongK1rwo8do465EkAyZau221jwcS48gyEco4W13DKegS5o4S8wj899pogBxC0yU4e4812odUcEaEG2a7o88nxO0wogwWxK6UbU7C2O0UEbEkw9-1nwNgconxu2Cl0RwEAghyES4UUMW2y2i2i4EbE-223nwjVUhwEwTxC',
        '__sjsp': 'gtA1WwhwBq8I8MZbblA_35ADksO4dcjfs9NsqVbPj8bcOsdE2xmWEJEmMGlO8F4y2eB5wy14ixsudg-9c54eg2qyi0344',
        '__comet_req': '7',
        'fb_dtsg': 'NAftM4s05XGPEdg-IbvSTtlHPXt9SPCmtXH-Sr1OBkVb3jsqC1em6Pw:17864863018060157:1769985092',
        'jazoest': '26200',
        'lsd': 'pD0B-QELbLVH3Ixu_JX8Cs',
        '__spin_r': '1035610417',
        '__spin_b': 'trunk',
        '__spin_t': '1774043853',
        '__crn': 'comet.igweb.PolarisProfilePostsTabRoute',
        'fb_api_caller_class': 'RelayModern',
        'fb_api_req_friendly_name': 'PolarisSearchBoxRefetchableQuery',
        'server_timestamps': 'true',
        'variables': query,
        'doc_id': '25877989431883913',
    }

    response = requests.post('https://www.instagram.com/graphql/query',
                             cookies=cookies, headers=headers, data=data)

    if response.status_code == 200:
        return response.json()


# d = search_by_name_instagram("hondamotosmaroc")
# print(d)


def get_profile_data(id: str):
    pass


def avito():
    import requests
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:148.0) Gecko/20100101 Firefox/148.0',
        'Accept': 'application/graphql-response+json,application/json;q=0.9',
        'Accept-Language': 'fr',
        # 'Accept-Encoding': 'gzip, deflate, br, zstd',
        'Referer': 'https://www.avito.ma/',
        'content-type': 'application/json',
        'visitorid': '36737bd7-8e29-4ba5-a24c-c097e73214bd',
        'user-session-id': 'undefined',
        'authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJodHRwczovL2FwaS5hdml0by5tYSIsImlhdCI6MTc3MzM1MDc3NSwiZXhwIjoxNzc1OTc4Nzc1LCJzdWIiOiI2OTQyMDYwIiwibmFtZSI6IkFtaW5lIiwiZW1haWwiOiJvcGVua29yZTkxQGdtYWlsLmNvbSIsInJvbGVzIjpbXSwic2Vzc2lvbklkIjoibWMxeDgwY2M4ODMwYzgzMDg1NWJhYzNjODZkZjQ3MWUxZGI3MDBjMTJkYTEiLCJhY2NvdW50SWQiOiI2OTQyMDYwIiwiY29tcGFueUFkIjoiMCJ9.Lul_B5Qm5IWvI0rE4go4VkUyiXye6iFDCGEhKAMO8UU',
        'Origin': 'https://www.avito.ma',
        'Sec-GPC': '1',
        'Connection': 'keep-alive',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-site',
    }

    json_data = {
        'operationName': 'getCombinedListingAds',
        'variables': {
            'premiumQuery': {
                'filters': {
                    'ad': {
                        'text': 'magasin',
                        'categoryId': 1060,
                        'type': 'SELL',
                        'location': {
                            'cityIds': [
                                12,
                            ],
                            'areaIds': [],
                        },
                        'params': {
                            'singleMatch': {
                                'text': [],
                                'numeric': [],
                                'boolean': [],
                            },
                            'listMatch': {
                                'textList': [],
                                'numericList': [],
                            },
                            'rangeMatch': [
                                {
                                    'name': 'size',
                                    'value': {
                                        'lessThanOrEqual': 80,
                                    },
                                },
                            ],
                        },
                        'isPremium': True,
                        'hasImage': True,
                    },
                    'extension': {
                        'includeNewConstructionAds': True,
                    },
                },
                'page': {
                    'number': 1,
                    'size': 3,
                },
            },
            'regularQuery': {
                'filters': {
                    'ad': {
                        'text': 'magasin',
                        'categoryId': 1060,
                        'type': 'SELL',
                        'location': {
                            'cityIds': [
                                12,
                            ],
                            'areaIds': [],
                        },
                        'params': {
                            'singleMatch': {
                                'text': [],
                                'numeric': [],
                                'boolean': [],
                            },
                            'listMatch': {
                                'textList': [],
                                'numericList': [],
                            },
                            'rangeMatch': [
                                {
                                    'name': 'size',
                                    'value': {
                                        'lessThanOrEqual': 80,
                                    },
                                },
                            ],
                        },
                    },
                    'extension': {
                        'includeNewConstructionAds': True,
                        'extendPublishedAdsSearchIfNeeded': True,
                    },
                },
                'page': {
                    'number': 1,
                    'size': 35,
                },
                'sort': {
                    'adProperty': 'SCORE',
                    'sortOrder': 'DESC',
                },
                'searchAlgorithm': 'KIT10X',
            },
        },
        'query': 'query getCombinedListingAds($premiumQuery: ListingAdsSearchQuery!, $regularQuery: ListingAdsSearchQuery!) {\n  premiumAds: getListingAds(query: $premiumQuery) {\n    ads {\n      details {\n        ... on PublishedAd {\n          adId\n          listId\n          category {\n            id\n            name\n            parent {\n              id\n              name\n              __typename\n            }\n            __typename\n          }\n          type {\n            key\n            name\n            __typename\n          }\n          title\n          description\n          price {\n            withCurrency\n            withoutCurrency\n            __typename\n          }\n          monthlyPayment {\n            withCurrency\n            withoutCurrency\n            __typename\n          }\n          oldPrice {\n            withCurrency\n            withoutCurrency\n            __typename\n          }\n          discount\n          params {\n            secondary {\n              ... on TextAdParam {\n                id\n                name\n                textValue\n                fullValue\n                trackingValue\n                __typename\n              }\n              ... on NumericAdParam {\n                id\n                name\n                numericValue\n                fullValue\n                unit\n                __typename\n              }\n              ... on BooleanAdParam {\n                id\n                name\n                booleanValue\n                __typename\n              }\n              __typename\n            }\n            extra {\n              ... on TextAdParam {\n                id\n                name\n                textValue\n                fullValue\n                trackingValue\n                __typename\n              }\n              ... on NumericAdParam {\n                id\n                name\n                numericValue\n                fullValue\n                unit\n                __typename\n              }\n              ... on BooleanAdParam {\n                id\n                name\n                booleanValue\n                __typename\n              }\n              __typename\n            }\n            __typename\n          }\n          media {\n            defaultImage {\n              paths {\n                standard\n                largeThumbnail\n                __typename\n              }\n              __typename\n            }\n            media {\n              images {\n                paths {\n                  standard\n                  largeThumbnail\n                  __typename\n                }\n                __typename\n              }\n              __typename\n            }\n            __typename\n          }\n          seller {\n            ... on PrivateProfile {\n              accountId\n              name\n              phone {\n                number\n                verified\n                __typename\n              }\n              __typename\n            }\n            ... on StoreProfile {\n              storeId\n              name\n              phone {\n                number\n                verified\n                __typename\n              }\n              logo {\n                defaultPath\n                __typename\n              }\n              __typename\n            }\n            __typename\n          }\n          sellerType\n          location {\n            city {\n              id\n              name\n              __typename\n            }\n            area {\n              id\n              name\n              __typename\n            }\n            address\n            __typename\n          }\n          listTime\n          isHighlighted\n          isInMyFavorites\n          offersShipping\n          isEcommerce\n          isImmoneuf\n          isPremium\n          isUrgent\n          isHotDeal\n          __typename\n        }\n        __typename\n      }\n      __typename\n    }\n    __typename\n  }\n  regularAds: getListingAds(query: $regularQuery) {\n    count {\n      total\n      __typename\n    }\n    ads {\n      details {\n        ... on PublishedAd {\n          adId\n          listId\n          category {\n            id\n            name\n            parent {\n              id\n              name\n              __typename\n            }\n            __typename\n          }\n          type {\n            key\n            name\n            __typename\n          }\n          title\n          description\n          price {\n            withCurrency\n            withoutCurrency\n            __typename\n          }\n          monthlyPayment {\n            withCurrency\n            withoutCurrency\n            __typename\n          }\n          oldPrice {\n            withCurrency\n            withoutCurrency\n            __typename\n          }\n          discount\n          params {\n            secondary {\n              ... on TextAdParam {\n                id\n                name\n                textValue\n                fullValue\n                trackingValue\n                __typename\n              }\n              ... on NumericAdParam {\n                id\n                name\n                numericValue\n                fullValue\n                unit\n                __typename\n              }\n              ... on BooleanAdParam {\n                id\n                name\n                booleanValue\n                __typename\n              }\n              __typename\n            }\n            extra {\n              ... on TextAdParam {\n                id\n                name\n                textValue\n                fullValue\n                trackingValue\n                __typename\n              }\n              ... on NumericAdParam {\n                id\n                name\n                numericValue\n                fullValue\n                unit\n                __typename\n              }\n              ... on BooleanAdParam {\n                id\n                name\n                booleanValue\n                __typename\n              }\n              __typename\n            }\n            __typename\n          }\n          media {\n            defaultImage {\n              paths {\n                standard\n                largeThumbnail\n                __typename\n              }\n              __typename\n            }\n            media {\n              images {\n                paths {\n                  standard\n                  largeThumbnail\n                  __typename\n                }\n                __typename\n              }\n              videos {\n                defaultPath\n                __typename\n              }\n              __typename\n            }\n            __typename\n          }\n          seller {\n            ... on PrivateProfile {\n              accountId\n              name\n              phone {\n                number\n                verified\n                __typename\n              }\n              __typename\n            }\n            ... on StoreProfile {\n              storeId\n              name\n              phone {\n                number\n                verified\n                __typename\n              }\n              logo {\n                defaultPath\n                __typename\n              }\n              isVerifiedSeller\n              __typename\n            }\n            __typename\n          }\n          sellerType\n          location {\n            city {\n              id\n              name\n              __typename\n            }\n            area {\n              id\n              name\n              __typename\n            }\n            address\n            __typename\n          }\n          listTime\n          isHighlighted\n          isInMyFavorites\n          offersShipping\n          isEcommerce\n          isImmoneuf\n          isUrgent\n          isHotDeal\n          __typename\n        }\n        ... on NewConstructionAd {\n          title\n          price {\n            withCurrency\n            withoutCurrency\n            __typename\n          }\n          location {\n            city {\n              id\n              name\n              __typename\n            }\n            area {\n              id\n              name\n              __typename\n            }\n            address\n            __typename\n          }\n          rooms\n          bathrooms\n          size\n          externalLink\n          media {\n            defaultImage {\n              defaultPath\n              __typename\n            }\n            __typename\n          }\n          __typename\n        }\n        __typename\n      }\n      searchExtension {\n        extensionType\n        __typename\n      }\n      __typename\n    }\n    __typename\n  }\n}',
    }

    response = requests.post(
        'https://gateway.avito.ma/graphql', headers=headers, json=json_data)
    if response.status_code == 200:
        print(response.json())


# avito()

# Note: json_data will not be serialized by requests
# exactly as it was in the original request.
# data = '{"operationName":"getCombinedListingAds","variables":{"premiumQuery":{"filters":{"ad":{"text":"magasin","categoryId":1060,"type":"SELL","location":{"cityIds":[12],"areaIds":[]},"params":{"singleMatch":{"text":[],"numeric":[],"boolean":[]},"listMatch":{"textList":[],"numericList":[]},"rangeMatch":[{"name":"size","value":{"lessThanOrEqual":80}}]},"isPremium":true,"hasImage":true},"extension":{"includeNewConstructionAds":true}},"page":{"number":1,"size":3}},"regularQuery":{"filters":{"ad":{"text":"magasin","categoryId":1060,"type":"SELL","location":{"cityIds":[12],"areaIds":[]},"params":{"singleMatch":{"text":[],"numeric":[],"boolean":[]},"listMatch":{"textList":[],"numericList":[]},"rangeMatch":[{"name":"size","value":{"lessThanOrEqual":80}}]}},"extension":{"includeNewConstructionAds":true,"extendPublishedAdsSearchIfNeeded":true}},"page":{"number":1,"size":35},"sort":{"adProperty":"SCORE","sortOrder":"DESC"},"searchAlgorithm":"KIT10X"}},"query":"query getCombinedListingAds($premiumQuery: ListingAdsSearchQuery!, $regularQuery: ListingAdsSearchQuery!) {\\n  premiumAds: getListingAds(query: $premiumQuery) {\\n    ads {\\n      details {\\n        ... on PublishedAd {\\n          adId\\n          listId\\n          category {\\n            id\\n            name\\n            parent {\\n              id\\n              name\\n              __typename\\n            }\\n            __typename\\n          }\\n          type {\\n            key\\n            name\\n            __typename\\n          }\\n          title\\n          description\\n          price {\\n            withCurrency\\n            withoutCurrency\\n            __typename\\n          }\\n          monthlyPayment {\\n            withCurrency\\n            withoutCurrency\\n            __typename\\n          }\\n          oldPrice {\\n            withCurrency\\n            withoutCurrency\\n            __typename\\n          }\\n          discount\\n          params {\\n            secondary {\\n              ... on TextAdParam {\\n                id\\n                name\\n                textValue\\n                fullValue\\n                trackingValue\\n                __typename\\n              }\\n              ... on NumericAdParam {\\n                id\\n                name\\n                numericValue\\n                fullValue\\n                unit\\n                __typename\\n              }\\n              ... on BooleanAdParam {\\n                id\\n                name\\n                booleanValue\\n                __typename\\n              }\\n              __typename\\n            }\\n            extra {\\n              ... on TextAdParam {\\n                id\\n                name\\n                textValue\\n                fullValue\\n                trackingValue\\n                __typename\\n              }\\n              ... on NumericAdParam {\\n                id\\n                name\\n                numericValue\\n                fullValue\\n                unit\\n                __typename\\n              }\\n              ... on BooleanAdParam {\\n                id\\n                name\\n                booleanValue\\n                __typename\\n              }\\n              __typename\\n            }\\n            __typename\\n          }\\n          media {\\n            defaultImage {\\n              paths {\\n                standard\\n                largeThumbnail\\n                __typename\\n              }\\n              __typename\\n            }\\n            media {\\n              images {\\n                paths {\\n                  standard\\n                  largeThumbnail\\n                  __typename\\n                }\\n                __typename\\n              }\\n              __typename\\n            }\\n            __typename\\n          }\\n          seller {\\n            ... on PrivateProfile {\\n              accountId\\n              name\\n              phone {\\n                number\\n                verified\\n                __typename\\n              }\\n              __typename\\n            }\\n            ... on StoreProfile {\\n              storeId\\n              name\\n              phone {\\n                number\\n                verified\\n                __typename\\n              }\\n              logo {\\n                defaultPath\\n                __typename\\n              }\\n              __typename\\n            }\\n            __typename\\n          }\\n          sellerType\\n          location {\\n            city {\\n              id\\n              name\\n              __typename\\n            }\\n            area {\\n              id\\n              name\\n              __typename\\n            }\\n            address\\n            __typename\\n          }\\n          listTime\\n          isHighlighted\\n          isInMyFavorites\\n          offersShipping\\n          isEcommerce\\n          isImmoneuf\\n          isPremium\\n          isUrgent\\n          isHotDeal\\n          __typename\\n        }\\n        __typename\\n      }\\n      __typename\\n    }\\n    __typename\\n  }\\n  regularAds: getListingAds(query: $regularQuery) {\\n    count {\\n      total\\n      __typename\\n    }\\n    ads {\\n      details {\\n        ... on PublishedAd {\\n          adId\\n          listId\\n          category {\\n            id\\n            name\\n            parent {\\n              id\\n              name\\n              __typename\\n            }\\n            __typename\\n          }\\n          type {\\n            key\\n            name\\n            __typename\\n          }\\n          title\\n          description\\n          price {\\n            withCurrency\\n            withoutCurrency\\n            __typename\\n          }\\n          monthlyPayment {\\n            withCurrency\\n            withoutCurrency\\n            __typename\\n          }\\n          oldPrice {\\n            withCurrency\\n            withoutCurrency\\n            __typename\\n          }\\n          discount\\n          params {\\n            secondary {\\n              ... on TextAdParam {\\n                id\\n                name\\n                textValue\\n                fullValue\\n                trackingValue\\n                __typename\\n              }\\n              ... on NumericAdParam {\\n                id\\n                name\\n                numericValue\\n                fullValue\\n                unit\\n                __typename\\n              }\\n              ... on BooleanAdParam {\\n                id\\n                name\\n                booleanValue\\n                __typename\\n              }\\n              __typename\\n            }\\n            extra {\\n              ... on TextAdParam {\\n                id\\n                name\\n                textValue\\n                fullValue\\n                trackingValue\\n                __typename\\n              }\\n              ... on NumericAdParam {\\n                id\\n                name\\n                numericValue\\n                fullValue\\n                unit\\n                __typename\\n              }\\n              ... on BooleanAdParam {\\n                id\\n                name\\n                booleanValue\\n                __typename\\n              }\\n              __typename\\n            }\\n            __typename\\n          }\\n          media {\\n            defaultImage {\\n              paths {\\n                standard\\n                largeThumbnail\\n                __typename\\n              }\\n              __typename\\n            }\\n            media {\\n              images {\\n                paths {\\n                  standard\\n                  largeThumbnail\\n                  __typename\\n                }\\n                __typename\\n              }\\n              videos {\\n                defaultPath\\n                __typename\\n              }\\n              __typename\\n            }\\n            __typename\\n          }\\n          seller {\\n            ... on PrivateProfile {\\n              accountId\\n              name\\n              phone {\\n                number\\n                verified\\n                __typename\\n              }\\n              __typename\\n            }\\n            ... on StoreProfile {\\n              storeId\\n              name\\n              phone {\\n                number\\n                verified\\n                __typename\\n              }\\n              logo {\\n                defaultPath\\n                __typename\\n              }\\n              isVerifiedSeller\\n              __typename\\n            }\\n            __typename\\n          }\\n          sellerType\\n          location {\\n            city {\\n              id\\n              name\\n              __typename\\n            }\\n            area {\\n              id\\n              name\\n              __typename\\n            }\\n            address\\n            __typename\\n          }\\n          listTime\\n          isHighlighted\\n          isInMyFavorites\\n          offersShipping\\n          isEcommerce\\n          isImmoneuf\\n          isUrgent\\n          isHotDeal\\n          __typename\\n        }\\n        ... on NewConstructionAd {\\n          title\\n          price {\\n            withCurrency\\n            withoutCurrency\\n            __typename\\n          }\\n          location {\\n            city {\\n              id\\n              name\\n              __typename\\n            }\\n            area {\\n              id\\n              name\\n              __typename\\n            }\\n            address\\n            __typename\\n          }\\n          rooms\\n          bathrooms\\n          size\\n          externalLink\\n          media {\\n            defaultImage {\\n              defaultPath\\n              __typename\\n            }\\n            __typename\\n          }\\n          __typename\\n        }\\n        __typename\\n      }\\n      searchExtension {\\n        extensionType\\n        __typename\\n      }\\n      __typename\\n    }\\n    __typename\\n  }\\n}"}'
# response = requests.post('https://gateway.avito.ma/graphql', headers=headers, data=data)

# with open("GetUnitList.json", "r") as f:
#    d = json.load(f)
#    p = json.loads(d)
# with open("fin.json", "w+") as ff:
#    json.dump(p, ff, indent=4)
#
# pip install requests
#
# username = 'openkore91_dpDF4'
# password = '~pUldq8+_X8j'
# proxy = 'dc.oxylabs.io:8000'
#
# proxies = {
#    "https": ('https://user-%s:%s@%s' % (username, password, proxy))
# }
#
# response = requests.get("https://ip.oxylabs.io/location", proxies=proxies)
#
# print(response.content)
#


# download_path = os.path.join(os.getcwd(), "download")
# prefs = {
#    "download.default_directory": download_path,
#    "plugins.always_open_pdf_externally": True
# }
# driver = setup_driver(True, prefs)
url = "https://investorhub.financialexpress.net/Pdf/fdd/en-gb/fdd/BUH7/U/?specialunittype=ORDN&priipproductcode="
# driver.get(url)


def download_pdf(url: str, name: str) -> int:
    response = fetch_with_backoff(url)
    if response:
        with open(f"{name}.pdf", "wb") as f:
            f.write(response.content)
            return 0
    return 1


def missing_funds() -> list[dict]:
    path = get_xlsx_filepath("financial_discount.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["Funds"]

    data = []
    i = 2
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        f = {}
        if row[1].value:
            i += 1
            continue
        f["name"] = row[0].value
        f["isin"] = row[1].value
        f["url"] = row[2].value
        f["row_index"] = i
        i += 1
        data.append(f)
    return data


def retry_missing_funds():
    data = missing_funds()
    new_data = []
    new_fund = {}
    # project_root = Path(__file__).resolve().parent.parent
    project_root = Path(__file__).resolve().parent
    i = 0
    for fund in data:
        print(f"downloading [{i+1}/{len(data)}]")
        p = join(project_root, "download", f"{fund['row_index']}")
        err = download_pdf(fund["url"], p)
        sleep(uniform(2, 5))
        new_fund.update(fund)
        new_fund.update({"error": err})
        new_data.append(new_fund)
        i += 1
    return new_data


# data = retry_missing_funds()

# utils.write_json("financial_missing.json", data)


def get_pdf_files(folder_name):
    folder_path = Path(folder_name)
    if not folder_path.exists():
        raise FileNotFoundError(f"Folder '{folder_name}' does not exist")
    if not folder_path.is_dir():
        raise NotADirectoryError(f"'{folder_name}' is not a directory")
    pdf_files = []
    for file in folder_path.iterdir():
        if file.is_file() and file.suffix.lower() == '.pdf':
            pdf_files.append(file.name)
    return pdf_files


def isin_from_text(text: str) -> str:
    isin_pattern = r"[A-Z]{2}[A-Z0-9]{9}[0-9]"
    isin = findall(isin_pattern, text)
    if len(isin) > 0:
        return isin[0]
    return ""


def isin_from_pdf(file: str) -> str | None:
    try:
        reader = PdfReader(file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
        return isin_from_text(text)
    except EmptyFileError:
        return None


def get_missing_isins():
    data = get_pdf_files("download")
    isins = []
    i = 0
    for pdf in data:
        print(f"[{i+1}/{len(data)}] reading {pdf} ")
        p = join("download", pdf)
        index = int(pdf[:len(pdf) - 4])
        isin = isin_from_pdf(p)
        isins.append({"index": index, "filename": pdf, "isin": isin})
        i += 1
    utils.write_json("financial_isins.json", isins)
    fix_missing_isins(isins)


def fix_missing_isins(data: list[dict]):
    path = utils.get_xlsx_filepath("financial_discount.xlsx")
    wb = load_workbook(path)
    ws = wb["Funds"]
    for fund in data:
        ws.cell(fund["index"], 2, fund["isin"])

    wb.save(path)
    wb.close()
    print(path)
    return


def count_empty_isins(filename: str, sheet: str) -> int:
    path = utils.get_xlsx_filepath(filename)
    wb = load_workbook(path)
    ws = wb[sheet]
    empty_count = 0
    for row in ws.iter_rows(min_row=1, values_only=False):
        cell_b = row[1]  # Column B is index 1
        if cell_b.value is None or str(cell_b.value).strip() == '':
            empty_count += 1
    wb.close()
    return empty_count
